from __future__ import annotations

import csv
import math
import os
from datetime import datetime
from typing import Callable, Dict, List, Optional, Sequence, Tuple

import numpy as np

from .models import (
    MULTI_START_SCALAR_FIELDS,
    PROTECTIONS,
    SCALAR_FIELDS,
    SINGLE_START_COLUMN_MAP,
    SINGLE_START_SCALAR_MAP,
    START_TYPES,
    StartupDataset,
    StartupRecord,
    StartupSeries,
)
from .startup_detection import detect_csv_type, extract_multi_start_rows, format_datetime, normalize_header


ProgressCallback = Callable[[float, str], None]


def read_csv_rows(file_path: str, progress_callback: ProgressCallback | None = None) -> List[List[str]]:
    extension = os.path.splitext(file_path)[1].lower()
    if extension == ".xlsx":
        return _read_xlsx_rows(file_path, progress_callback=progress_callback)

    raw_bytes = _read_csv_bytes(file_path, progress_callback=progress_callback)
    last_rows: List[List[str]] = []
    for encoding_name in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw_bytes.decode(encoding_name)
        except UnicodeDecodeError:
            continue
        if progress_callback:
            progress_callback(0.62, f"Decodificando CSV ({encoding_name})...")
        rows = _parse_csv_text(text, progress_callback=progress_callback)
        if rows:
            return rows
        last_rows = rows

    text = raw_bytes.decode("utf-8", errors="replace")
    return _parse_csv_text(text, progress_callback=progress_callback) or last_rows


def _read_csv_bytes(file_path: str, progress_callback: ProgressCallback | None = None) -> bytes:
    total_size = max(1, math.ceil(float(max(1, os.path.getsize(file_path)))))
    chunk_size = 1024 * 1024
    parts: list[bytes] = []
    read_bytes = 0
    with open(file_path, "rb") as csv_file:
        while True:
            chunk = csv_file.read(chunk_size)
            if not chunk:
                break
            parts.append(chunk)
            read_bytes += len(chunk)
            if progress_callback:
                progress_callback(min(0.55, 0.55 * (read_bytes / total_size)), "Leyendo archivo CSV...")
    return b"".join(parts).replace(b"\x00", b"")


def _parse_csv_text(text: str, progress_callback: ProgressCallback | None = None) -> List[List[str]]:
    lines = text.splitlines()
    try:
        rows: List[List[str]] = []
        total_lines = max(1, len(lines))
        for index, row in enumerate(csv.reader(lines), start=1):
            rows.append(row)
            if progress_callback and (index == 1 or index % 250 == 0 or index == total_lines):
                progress = 0.62 + 0.28 * (index / total_lines)
                progress_callback(min(0.90, progress), f"Parseando filas del CSV ({index}/{total_lines})...")
        return rows
    except csv.Error:
        return []


def _read_xlsx_rows(file_path: str, progress_callback: ProgressCallback | None = None) -> List[List[str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError(f"No se pudo abrir XLSX porque falta openpyxl: {exc}") from exc

    if progress_callback:
        progress_callback(0.15, "Abriendo libro XLSX...")
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        total_rows = max(1, int(getattr(worksheet, "max_row", 0) or 1))
        rows: List[List[str]] = []
        for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            rows.append(["" if value is None else str(value) for value in row])
            if progress_callback and (index == 1 or index % 100 == 0 or index == total_rows):
                progress = 0.20 + 0.70 * (index / total_rows)
                progress_callback(min(0.90, progress), f"Leyendo filas del XLSX ({index}/{total_rows})...")
        return rows
    finally:
        workbook.close()


def parse_numeric(value, factor: float = 1.0, default: float = 0.0) -> float:
    try:
        if value is None or not str(value).strip():
            return default
        return float(value) * factor
    except Exception:
        return default


def to_float_or_nan(value) -> float:
    try:
        if value is None or str(value).strip() == "":
            return np.nan
        return float(value)
    except Exception:
        return np.nan


def try_parse_datetime_text(value) -> Optional[datetime]:
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    if "T" in text:
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            pass
    return None


def build_empty_series() -> StartupSeries:
    return StartupSeries()


def validate_rows(rows: Sequence[Sequence[str]]) -> List[str]:
    issues: List[str] = []
    if not rows:
        issues.append("El CSV está vacío.")
        return issues
    if not any(any(str(cell).strip() for cell in row) for row in rows):
        issues.append("El CSV no contiene datos útiles.")
    csv_type = detect_csv_type(rows)
    if csv_type is None:
        issues.append("No se pudo identificar el formato del CSV.")
    return issues


def parse_scalars_from_multi_row(row: Sequence[str]) -> Dict[str, float]:
    scalars: Dict[str, float] = {}
    index = 21
    for name, factor in MULTI_START_SCALAR_FIELDS:
        raw_value = row[index] if index < len(row) else None
        scalars[name] = parse_numeric(raw_value, factor)
        index += 1
    return scalars


def _convert_single_scalar(raw_value, factor: float = 1.0, offset: float = 0.0) -> float:
    base = to_float_or_nan(raw_value)
    if np.isnan(base):
        return np.nan
    return base * factor + offset


def resolve_header_index(header_map: Dict[str, int], *candidate_names: str) -> Optional[int]:
    normalized_candidates = [normalize_header(name) for name in candidate_names if normalize_header(name)]
    for candidate in normalized_candidates:
        if candidate in header_map:
            return header_map[candidate]
    for candidate in normalized_candidates:
        for header_name, index in header_map.items():
            if candidate in header_name or header_name in candidate:
                return index
    return None


def _safe_row_value(row: Sequence[str], index: Optional[int], default: str = "") -> str:
    if index is None or index >= len(row):
        return default
    return str(row[index]).strip()


def parse_multi_start_record(row: Sequence[str], label: str, dt_tuple: Tuple[int, int, int, int, int, int]) -> StartupRecord:
    protection = PROTECTIONS.get(int(parse_numeric(row[0] if len(row) > 0 else None)), "Unknown")
    start_type = START_TYPES.get(int(parse_numeric(row[1] if len(row) > 1 else None)), "Unknown")
    timestamp_text = format_datetime(dt_tuple)
    description = "Sin descripción" if "15163" in row[5:21] else " ".join(cell for cell in row[5:21] if str(cell).strip())
    scalars = {
        "Fecha completa": timestamp_text,
        "Descripción": description or "Sin descripción",
        "Protección": protection,
        "Tipo arranque": start_type,
    }
    scalars.update(parse_scalars_from_multi_row(row))

    index = 21 + len(MULTI_START_SCALAR_FIELDS) + 5

    def read_block(size: int, factor: float) -> np.ndarray:
        nonlocal index
        values = [parse_numeric(row[index + offset] if index + offset < len(row) else None, factor) for offset in range(size)]
        index += size
        return np.array(values, dtype=float)

    speed = read_block(300, 0.1)
    current = read_block(300, 0.1)
    torque = read_block(300, 0.1)
    load_torque = read_block(180, 0.1)
    motor_torque = read_block(180, 0.1)

    raw_dual = list(row[index:index + 300])
    dual_current = np.array([], dtype=float)
    if any(int(cell) != 16383 for cell in raw_dual[1:] if str(cell).strip()):
        dual_current = np.array([parse_numeric(cell, 0.1) for cell in raw_dual], dtype=float)
    index += 300

    harmonic_amp = []
    harmonic_freq_raw = []
    for _ in range(10):
        harmonic_amp.append(parse_numeric(row[index] if index < len(row) else None))
        index += 1
        harmonic_freq_raw.append(parse_numeric(row[index] if index < len(row) else None))
        index += 1

    harmonic_amp_arr = np.array(harmonic_amp, dtype=float) / math.sqrt(2)
    harmonic_freq_raw_arr = np.array(harmonic_freq_raw, dtype=float)
    harmonic_freq_hz_arr = harmonic_freq_raw_arr / 0.5632

    sample_rate = scalars.get("Sampling rate (ms)", 0.0)
    time_axis = np.arange(len(speed)) * (sample_rate / 1000.0 if sample_rate else 0.0)

    return StartupRecord(
        label=label,
        timestamp_text=timestamp_text,
        timestamp_dt=datetime(*dt_tuple),
        protection=protection,
        start_type=start_type,
        description=description or "Sin descripción",
        scalars=scalars,
        series=StartupSeries(
            speed=speed,
            current=current,
            torque=torque,
            time=time_axis,
            load_torque=load_torque,
            motor_torque=motor_torque,
            dual_current=dual_current,
            harmonic_amp=harmonic_amp_arr,
            harmonic_freq_raw=harmonic_freq_raw_arr,
            harmonic_freq_hz=harmonic_freq_hz_arr,
        ),
        source_kind="multi_file",
    )


def parse_multi_start_csv(rows: Sequence[Sequence[str]]) -> List[StartupRecord]:
    items = extract_multi_start_rows(rows)
    return [parse_multi_start_record(row, label, dt_tuple) for dt_tuple, label, row in items]


def parse_single_start_csv(rows: Sequence[Sequence[str]]) -> List[StartupRecord]:
    if len(rows) < 2:
        return []

    header_map = {normalize_header(name): index for index, name in enumerate(rows[0])}
    meta_row = rows[1]
    scalars = {name: np.nan for name in SCALAR_FIELDS}

    for raw_name, (target_name, factor, offset) in SINGLE_START_SCALAR_MAP.items():
        index = resolve_header_index(header_map, raw_name, target_name)
        if index is not None and index < len(meta_row):
            scalars[target_name] = _convert_single_scalar(meta_row[index], factor, offset)

    timestamp_index = resolve_header_index(header_map, "time")
    protection_index = resolve_header_index(header_map, "protection")
    start_type_index = resolve_header_index(header_map, "typeofstart", "tipoarranque")
    description_index = resolve_header_index(header_map, "description", "descripción", "descripcion")

    timestamp_text = _safe_row_value(meta_row, timestamp_index, "Sin fecha")
    timestamp_dt = try_parse_datetime_text(timestamp_text)
    label = timestamp_text if timestamp_dt is None else timestamp_dt.strftime("%Y/%m/%d %H:%M:%S")
    protection = _safe_row_value(meta_row, protection_index, "Unknown")
    start_type = _safe_row_value(meta_row, start_type_index, "Unknown")
    description = _safe_row_value(meta_row, description_index, "Sin descripción") or "Sin descripción"

    enriched_scalars: Dict[str, float] = {
        "Fecha completa": label,
        "Descripción": description,
        "Protección": protection,
        "Tipo arranque": start_type,
    }
    enriched_scalars.update(scalars)

    series_columns = {
        "speed": [],
        "current": [],
        "torque": [],
        "time": [],
        "load_torque": [],
        "motor_torque": [],
        "dual_current": [],
        "harmonic_amp": [],
    }
    harmonic_rows: List[float] = []

    for row in rows[1:]:
        for raw_name, target in SINGLE_START_COLUMN_MAP.items():
            index = resolve_header_index(header_map, raw_name, target)
            if index is None or index >= len(row):
                continue
            value = to_float_or_nan(row[index])
            if target == "harmonic_amp_arms":
                series_columns["harmonic_amp"].append(value)
            elif target == "harmonic_freq":
                harmonic_rows.append(value)
            elif target in series_columns:
                series_columns[target].append(value)

    series = build_empty_series()
    for key in ("speed", "current", "torque", "load_torque", "motor_torque", "dual_current"):
        setattr(series, key, np.array(series_columns[key], dtype=float))

    if series_columns["time"]:
        series.time = np.array(series_columns["time"], dtype=float)
    else:
        sample_rate = enriched_scalars.get("Sampling rate (ms)", np.nan)
        step = sample_rate / 1000.0 if not np.isnan(sample_rate) else 0.0
        series.time = np.arange(len(series.speed)) * step

    series.harmonic_amp = np.array(series_columns["harmonic_amp"], dtype=float)
    series.harmonic_freq_hz = np.array(harmonic_rows, dtype=float)
    series.harmonic_freq_raw = np.array(harmonic_rows, dtype=float) * 0.5632

    return [
        StartupRecord(
            label=label,
            timestamp_text=label,
            timestamp_dt=timestamp_dt,
            protection=protection,
            start_type=start_type,
            description=description,
            scalars=enriched_scalars,
            series=series,
            source_kind="single_file",
        )
    ]


def parse_csv_dataset(rows: Sequence[Sequence[str]]) -> StartupDataset:
    issues = validate_rows(rows)
    csv_type = detect_csv_type(rows)
    if csv_type == "single_file":
        records = parse_single_start_csv(rows)
    elif csv_type == "multi_file":
        records = parse_multi_start_csv(rows)
    else:
        records = []
    view_mode = "multi_startup_view" if len(records) > 1 else "single_startup_view"
    if not records and not issues:
        issues.append("No se detectaron arranques válidos en el CSV.")
    return StartupDataset(csv_type=csv_type, view_mode=view_mode, records=records, validation_issues=issues)


def parse_csv_records_to_legacy(rows: Sequence[Sequence[str]]):
    dataset = parse_csv_dataset(rows)
    return dataset.csv_type, dataset.view_mode, dataset.to_legacy_records(), list(dataset.validation_issues)
